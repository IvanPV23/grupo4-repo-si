"""
utils/excel_acumulativo.py
Gestiona el reporte Excel persistente del sistema.
Agrega nuevas filas sin sobreescribir el histórico.
"""

import os
from datetime import datetime

# openpyxl es la librería estándar para .xlsx
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# Ruta del reporte acumulativo
RUTA_REPORTE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "data", "outputs", "reporte_acumulativo.xlsx"
)


def _asegurar_permisos(ruta: str):
    """Trata de hacer el archivo escribible si existe.

    En Windows puede que esté marcado como solo-lectura o usado por otro
    proceso; en Linux/cont. se intenta chmod 666.
    """
    try:
        if os.path.exists(ruta):
            # intento chmod; si falla en Windows, no pasa nada
            os.chmod(ruta, 0o666)
    except Exception as ex:
        print(f"[Excel] no se pudieron ajustar permisos de {ruta}: {ex}")


# Asegurarse de que la carpeta de salida exista y sea escribible
_output_dir = os.path.dirname(RUTA_REPORTE)
try:
    os.makedirs(_output_dir, exist_ok=True)
    # intentar dar permisos amplios para evitar errores en contenedores
    try:
        os.chmod(_output_dir, 0o777)
    except Exception:
        pass
except Exception as e:
    print(f"[Excel] no pudo preparar directorio de salida {_output_dir}: {e}")

COLUMNAS = [
    "ticket_id",
    "resumen",
    "tipo_incidencia",
    "tipo_atencion_sd",
    "area",
    "producto",
    "aplicativo",
    "informador",
    "urgencia_detectada",
    "tiempo_estimado_horas",
    "categoria_tiempo",
    "complejidad",
    "score_complejidad",
    "nivel_asignado",
    "mesa_asignada",
    "via_historico",
    "resultado",
    "razonamiento",
    "procesado_en",
]


def _crear_libro_nuevo(ruta: str):
    """Crea un Excel nuevo con el encabezado formateado."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Derivaciones"

    # Estilo de encabezado
    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(COLUMNAS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    # Anchos de columna razonables
    anchos = {
        "ticket_id": 14, "resumen": 45, "tipo_incidencia": 16,
        "tipo_atencion_sd": 28, "area": 18, "producto": 14,
        "aplicativo": 18, "informador": 22, "urgencia_detectada": 18,
        "tiempo_estimado_horas": 22, "complejidad": 14, "score_complejidad": 18,
        "nivel_asignado": 14, "mesa_asignada": 28, "via_historico": 14,
        "resultado": 28, "razonamiento": 60, "procesado_en": 20,
    }
    for col_idx, col_name in enumerate(COLUMNAS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = anchos.get(col_name, 18)

    try:
        os.makedirs(os.path.dirname(ruta), exist_ok=True)
        wb.save(ruta)
    except PermissionError as pe:
        # No se puede crear/guardar el archivo por permisos; loguear y continuar
        print(f"[Excel] permiso denegado al crear {ruta}: {pe}")
    except Exception as ex:
        print(f"[Excel] error al crear libro nuevo en {ruta}: {ex}")
    return wb


def agregar_fila_reporte(datos: dict) -> str:
    """
    Agrega una nueva fila al reporte acumulativo Excel.

    Args:
        datos: dict con claves iguales a COLUMNAS (campos faltantes quedan vacíos).

    Returns:
        Ruta del archivo Excel.
    """
    if not OPENPYXL_OK:
        print("[Excel] openpyxl no disponible — saltando escritura")
        return RUTA_REPORTE

    # Cargar o crear el libro
    if os.path.exists(RUTA_REPORTE):
        wb = openpyxl.load_workbook(RUTA_REPORTE)
        ws = wb.active
        # Verificar que la fila de encabezado esté sincronizada con COLUMNAS.
        # Si el archivo existente no tenía la nueva columna `categoria_tiempo`
        # (u otras que se agreguen en el futuro), la primera fila se corregirá
        # para evitar que las entradas subsiguientes queden desfasadas.
        existing_headers = [str(ws.cell(row=1, column=i).value).strip()
                            if ws.cell(row=1, column=i).value is not None else ''
                            for i in range(1, ws.max_column + 1)]
        # comprobar si el encabezado actual falta alguna de las columnas definidas
        # y, en tal caso, añadir una columna vacía en la posición adecuada.
        for idx, col_name in enumerate(COLUMNAS, start=1):
            if idx > len(existing_headers) or existing_headers[idx-1] != col_name:
                # falta columna en la hoja -> insertar
                ws.insert_cols(idx)
                # opcional: dejar celdas vacías en las filas existentes
                existing_headers.insert(idx-1, col_name)
                # avanzar índices de headers para chequeo siguiente
        # tras insertar columnas necesarias, reescribimos la fila de encabezado
        for col_idx, col_name in enumerate(COLUMNAS, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
    else:
        wb = _crear_libro_nuevo(RUTA_REPORTE)
        ws = wb.active

    # Agregar datos en la siguiente fila disponible
    fila = ws.max_row + 1
    datos.setdefault("procesado_en", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    for col_idx, col_name in enumerate(COLUMNAS, start=1):
        valor = datos.get(col_name, "")
        ws.cell(row=fila, column=col_idx, value=valor)

    try:
        wb.save(RUTA_REPORTE)
        print(f"[Excel] Fila {fila} agregada → {RUTA_REPORTE}")
    except PermissionError as pe:
        print(f"[Excel] permiso denegado al guardar fila en {RUTA_REPORTE}: {pe}")
    except Exception as ex:
        print(f"[Excel] error al guardar fila {fila} en {RUTA_REPORTE}: {ex}")
    return RUTA_REPORTE


def obtener_resumen_reporte() -> dict:
    """Retorna estadísticas básicas del reporte acumulativo."""
    if not OPENPYXL_OK or not os.path.exists(RUTA_REPORTE):
        return {"total": 0, "archivo": RUTA_REPORTE, "existe": False}

    wb = openpyxl.load_workbook(RUTA_REPORTE, read_only=True)
    ws = wb.active
    total = ws.max_row - 1  # -1 por el encabezado
    wb.close()

    return {
        "total": max(0, total),
        "archivo": RUTA_REPORTE,
        "existe": True,
    }
